from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle


ROOT = Path(__file__).resolve().parents[1]
PUBLIC_RESOURCES_DIR = ROOT / "public" / "resources"
HANDOUT_DIR = PUBLIC_RESOURCES_DIR / "handouts"
EXAMPLE_DIR = PUBLIC_RESOURCES_DIR / "examples"
WORKBOOK_DIR = PUBLIC_RESOURCES_DIR / "workbooks"
TMP_PDF_DIR = ROOT / "tmp" / "pdfs" / "public-resources"
TMP_SHEET_DIR = ROOT / "tmp" / "spreadsheets" / "public-resources"

SOURCE_HANDOUTS = {
    "ki-register-handout-use-case-pass-anatomie.pdf": ROOT.parents[1]
    / "docs"
    / "kurs-multiplikator"
    / "kurs-A"
    / "handouts"
    / "A2_UseCase_Pass_Anatomie.pdf",
    "ki-register-handout-dsb-workflow-arbeitsblatt.pdf": ROOT.parents[1]
    / "docs"
    / "kurs-multiplikator"
    / "kurs-B"
    / "handouts"
    / "B2_DSB_Workflow_Arbeitsblatt.pdf",
    "ki-register-handout-ki-einfuehrung-zu-governance.pdf": ROOT.parents[1]
    / "docs"
    / "kurs-multiplikator"
    / "kurs-C"
    / "handouts"
    / "C1_KI_Einfuehrung_zu_Governance.pdf",
    "ki-register-handout-workshop-integration.pdf": ROOT.parents[1]
    / "docs"
    / "kurs-multiplikator"
    / "kurs-C"
    / "handouts"
    / "C2_Workshop_Integration_Template.pdf",
}


DARK = "111111"
MID = "6B7280"
LIGHT = "E5E7EB"
SOFT = "F7F7F7"
WHITE = "FFFFFF"
WARNING = "D1D5DB"


def ensure_dirs() -> None:
    for directory in (HANDOUT_DIR, EXAMPLE_DIR, WORKBOOK_DIR, TMP_PDF_DIR, TMP_SHEET_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def sync_handouts() -> None:
    for target_name, source_path in SOURCE_HANDOUTS.items():
        if not source_path.exists():
            raise FileNotFoundError(f"Missing handout source: {source_path}")
        copy2(source_path, HANDOUT_DIR / target_name)


def build_styles():
    styles = getSampleStyleSheet()
    return {
        "brand": ParagraphStyle(
            "ResourceBrand",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            uppercase=True,
            textColor=colors.white,
            backColor=colors.black,
            leftIndent=0,
            rightIndent=0,
            borderPadding=(2, 6, 2),
            spaceAfter=8,
        ),
        "title": ParagraphStyle(
            "ResourceTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=21,
            leading=26,
            textColor=colors.black,
            spaceAfter=5,
        ),
        "subtitle": ParagraphStyle(
            "ResourceSubtitle",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=15,
            textColor=colors.HexColor("#4b5563"),
            spaceAfter=12,
        ),
        "section": ParagraphStyle(
            "ResourceSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.black,
            spaceBefore=10,
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "ResourceBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=15,
            textColor=colors.black,
            spaceAfter=5,
        ),
        "bullet": ParagraphStyle(
            "ResourceBullet",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=15,
            textColor=colors.black,
            leftIndent=12,
            bulletIndent=0,
            spaceAfter=2,
        ),
        "small": ParagraphStyle(
            "ResourceSmall",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#6b7280"),
        ),
    }


def add_pdf_header(story, styles, title: str, subtitle: str, context_rows: list[tuple[str, str]]) -> None:
    story.append(Paragraph("KI Register", styles["brand"]))
    story.append(Paragraph(title, styles["title"]))
    story.append(Paragraph(subtitle, styles["subtitle"]))

    if context_rows:
        table = PdfTable(
            context_rows,
            colWidths=[42 * mm, 120 * mm],
            style=TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEADING", (0, 0), (-1, -1), 12),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.HexColor("#111111")),
                    ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            ),
        )
        story.append(table)
        story.append(Spacer(1, 7))


def add_paragraph_section(story, styles, heading: str, paragraphs: list[str]) -> None:
    story.append(Paragraph(heading, styles["section"]))
    for paragraph in paragraphs:
        story.append(Paragraph(paragraph, styles["body"]))


def add_bullet_section(story, styles, heading: str, items: list[str]) -> None:
    story.append(Paragraph(heading, styles["section"]))
    for item in items:
        story.append(Paragraph(item, styles["bullet"], bulletText="-"))


def add_table_section(story, styles, heading: str, rows: list[tuple[str, str]]) -> None:
    story.append(Paragraph(heading, styles["section"]))
    table = PdfTable(
        rows,
        colWidths=[52 * mm, 110 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#f7f7f7"), colors.white]),
                ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.HexColor("#111111")),
                ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEADING", (0, 0), (-1, -1), 12),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]
        ),
    )
    story.append(table)


def draw_pdf_footer(canvas, doc) -> None:
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#d1d5db"))
    canvas.line(doc.leftMargin, 11 * mm, doc.pagesize[0] - doc.rightMargin, 11 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawString(doc.leftMargin, 7.5 * mm, "KI Register - kiregister.com")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 7.5 * mm, f"Seite {canvas.getPageNumber()}")
    canvas.restoreState()


def create_pdf(
    path: Path,
    title: str,
    subtitle: str,
    context_rows: list[tuple[str, str]],
    sections: list[dict[str, object]],
) -> None:
    styles = build_styles()
    story = []
    add_pdf_header(story, styles, title, subtitle, context_rows)

    for section in sections:
        kind = section["kind"]
        if kind == "paragraphs":
            add_paragraph_section(story, styles, section["heading"], section["items"])
        elif kind == "bullets":
            add_bullet_section(story, styles, section["heading"], section["items"])
        elif kind == "table":
            add_table_section(story, styles, section["heading"], section["rows"])

    document = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    document.build(story, onFirstPage=draw_pdf_footer, onLaterPages=draw_pdf_footer)


def create_use_case_pass_pdf(path: Path) -> None:
    create_pdf(
        path,
        "Beispiel Use-Case Pass",
        "Ein vollstaendiges Referenzdokument fuer einen dokumentierten KI-Einsatzfall mit Review- und Evidenzkette.",
        [
            ("Organisation", "Nordstadt Automotive Systems GmbH"),
            ("Einsatzfall", "Automatisierte Eingangspruefung technischer Lieferantendokumente"),
            ("Owner Rolle", "Leitung Supplier Quality"),
            ("Review-Stand", "REVIEWED - export- und verify-faehig"),
            ("Primaerer Nutzen", "Vorpruefung vor manueller Freigabe"),
            ("Externe Anschlussfaehigkeit", "OEM-Anfragen, Audit, Supplier Review"),
        ],
        [
            {
                "kind": "paragraphs",
                "heading": "Management-Zusammenfassung",
                "items": [
                    "Der Einsatzfall klassifiziert eingehende technische Lieferantendokumente nach Vollstaendigkeit, Relevanz und erkennbarem Nacharbeitsbedarf. Er beschleunigt die Vorpruefung, ersetzt aber keine finale fachliche Freigabe.",
                    "Der Pass macht sichtbar, warum der Fall governance-faehig ist: Verantwortlichkeit, Datenbezug, Wirkungslogik, Kontrollmassnahmen und Review-Zyklus sind in einer einzigen, extern anschlussfaehigen Einheit zusammengefuehrt.",
                ],
            },
            {
                "kind": "table",
                "heading": "Kernfelder des Einsatzfalls",
                "rows": [
                    ("System", "Technical Document Classifier"),
                    ("Nutzungskontext", "Supplier Quality / Vorpruefung von OEM- und Lieferantendokumenten"),
                    ("Datenbezug", "Technische Zeichnungen, Lastenhefte, Lieferantendokumente"),
                    ("Entscheidungsnaehe", "Vorbereitung mit manueller Freigabe"),
                    ("AI-Act Einordnung", "Minimales Risiko mit dokumentationsrelevantem Review"),
                    ("Naechster Review-Termin", "30. September 2026"),
                ],
            },
            {
                "kind": "bullets",
                "heading": "Kontrollen und Betriebsregeln",
                "items": [
                    "Keine finale Freigabe ohne menschliche Sichtung und dokumentierte Verantwortungsrolle.",
                    "Eingangsquellen und Dokumententypen sind pro Lieferantenprozess definiert; spontane Freitextimporte sind nicht freigegeben.",
                    "Quartalsweiser Review von Datenbasis, False-Positive-Mustern und dokumentierten Ausnahmen.",
                    "Verify-Link und PDF-Export werden nur aus dem freigegebenen Registereintrag erzeugt.",
                ],
            },
            {
                "kind": "table",
                "heading": "Evidenzkette",
                "rows": [
                    ("Interne Richtlinie", "AI Usage Policy 2026 / Abschnitt Supplier Quality"),
                    ("Pruefprotokoll", "Review vom 14. April 2026 durch Governance Lead"),
                    ("Nachweisartefakt", "PDF-Export fuer OEM- und Audit-Anfragen"),
                    ("Externe Referenz", "Verify-Link zur Echtheitspruefung des Passes"),
                ],
            },
            {
                "kind": "paragraphs",
                "heading": "Warum dieses Beispiel belastbar ist",
                "items": [
                    "Das Beispiel beschreibt nicht nur das System, sondern den konkreten Einsatz. Genau deshalb kann ein Auditor nachvollziehen, was die KI wirklich tut und welche Grenze fuer menschliche Freigabe gilt.",
                    "Die Informationen sind so strukturiert, dass sie intern gepflegt und extern in Due-Diligence-, Lieferketten- oder Audit-Situationen wiederverwendet werden koennen, ohne dass ein zweites Parallel-Dokument entsteht.",
                ],
            },
        ],
    )


def create_supplier_response_pdf(path: Path) -> None:
    create_pdf(
        path,
        "Beispielantwort fuer OEM-Anfragen",
        "Formalisierte Supplier-Response zu einem einzelnen KI-Einsatzfall mit Scope, Aussagen und Evidenzbezug.",
        [
            ("Anfragender", "OEM Procurement & Supplier Assurance"),
            ("Antwortende Organisation", "Nordstadt Automotive Systems GmbH"),
            ("Einsatzfall", "Automatisierte Eingangspruefung technischer Lieferantendokumente"),
            ("Antwortverantwortung", "Head of Governance + Supplier Quality Lead"),
            ("Antwortdatum", "15. April 2026"),
            ("Belegstruktur", "Use-Case Pass, Review-Protokoll, Verify-Link"),
        ],
        [
            {
                "kind": "paragraphs",
                "heading": "Scope der Antwort",
                "items": [
                    "Diese Antwort bezieht sich ausschliesslich auf den dokumentierten Einsatzfall zur Vorpruefung technischer Lieferantendokumente. Sie ist keine pauschale Aussage ueber saemtliche KI-Nutzung des Unternehmens.",
                    "Der Einsatzfall wird intern als Use-Case Pass gefuehrt. Alle Angaben in diesem Dossier sind auf dieselbe Falldefinition, denselben Review-Stand und dieselbe Verantwortungsrolle bezogen.",
                ],
            },
            {
                "kind": "table",
                "heading": "Kernaussagen fuer den Auftraggeber",
                "rows": [
                    ("Zweck", "Vorsortierung und Priorisierung eingehender technischer Dokumente"),
                    ("Automatisierte Entscheidung", "Nein - finale Freigaben erfolgen manuell"),
                    ("Betroffene Daten", "Technische Dokumente und zugehoerige Lieferantenmetadaten"),
                    ("Externe Wirkung", "Indirekte Wirkung auf Beschaffungs- und Qualitaetsprozesse"),
                    ("Aktueller Status", "REVIEWED"),
                    ("Lieferbarer Nachweis", "PDF-Export und Verify-Link des zugehoerigen Passes"),
                ],
            },
            {
                "kind": "bullets",
                "heading": "Was wir gegenueber OEMs verbindlich aussagen koennen",
                "items": [
                    "Der Einsatzfall ist intern einer verantwortlichen Rolle zugeordnet und wird in einem festen Review-Zyklus gepflegt.",
                    "Es besteht keine vollautomatische Freigabe oder Zurueckweisung von Lieferantenunterlagen ohne menschliche Kontrolle.",
                    "Kontrollpunkte, Review-Status und externe Nachweise werden aus derselben Registerlogik erzeugt wie die interne Dokumentation.",
                    "Bekannte Einschraenkungen und offene Fragen werden nicht ausserhalb des Passes versteckt, sondern als Review-Themen sichtbar gehalten.",
                ],
            },
            {
                "kind": "table",
                "heading": "Beigefuegte oder referenzierte Evidenz",
                "rows": [
                    ("Use-Case Pass", "Interner Pass mit Owner, Zweck, Datenbezug und Risikoeinordnung"),
                    ("Review-Protokoll", "Pruefvermerk des letzten Governance-Reviews"),
                    ("Verify-Link", "Externe Echtheitspruefung des freigegebenen Artefakts"),
                    ("Naechster Review", "Terminierter Folge-Review im regulativen Betriebszyklus"),
                ],
            },
            {
                "kind": "paragraphs",
                "heading": "Grenzen dieser Antwort",
                "items": [
                    "Dieses Dossier ersetzt weder Vertragspruefung noch eine weitergehende technische Due Diligence. Es dient dazu, einen konkreten Einsatzfall belastbar und vergleichbar zu beantworten.",
                    "Wenn der Auftraggeber weitergehende Fragen stellt, werden diese nicht per E-Mail improvisiert beantwortet, sondern gegen denselben Pass und dieselbe Evidenzstruktur fortgeschrieben.",
                ],
            },
        ],
    )


def title_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=DARK)


def sub_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=SOFT)


def status_fill(value: str) -> PatternFill:
    mapping = {
        "UNREVIEWED": "F3F4F6",
        "REVIEW_RECOMMENDED": "E5E7EB",
        "REVIEWED": "D1D5DB",
        "PROOF_READY": "111111",
    }
    return PatternFill(fill_type="solid", fgColor=mapping.get(value, SOFT))


def status_font(value: str) -> Font:
    if value == "PROOF_READY":
        return Font(color=WHITE, bold=True)
    return Font(color=DARK, bold=True)


def set_sheet_base(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"


def style_title_block(worksheet, title: str, subtitle: str, width: int = 8) -> None:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    worksheet["A1"] = title
    worksheet["A2"] = subtitle
    worksheet["A1"].fill = title_fill()
    worksheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet["A2"].fill = sub_fill()
    worksheet["A2"].font = Font(color=DARK, size=10)
    worksheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 34


def style_section_row(worksheet, row_index: int, label: str, width: int = 8) -> None:
    worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=width)
    cell = worksheet.cell(row=row_index, column=1)
    cell.value = label
    cell.fill = sub_fill()
    cell.font = Font(color=DARK, bold=True, size=10)
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row_index].height = 20


def style_table_header(row) -> None:
    fill = title_fill()
    font = Font(color=WHITE, bold=True, size=10)
    border = Border(bottom=Side(style="thin", color=LIGHT))
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_body_cell(cell, bold: bool = False, fill: str | None = None) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color=LIGHT))
    cell.font = Font(color=DARK, bold=bold, size=10)
    if fill:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill)


def set_widths(worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def add_excel_table(worksheet, table_name: str, start_row: int, end_row: int, end_col: str) -> None:
    table = Table(displayName=table_name, ref=f"A{start_row}:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def create_use_case_register_workbook(path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    set_sheet_base(overview)
    style_title_block(
        overview,
        "Use-Case Register Review Matrix",
        "Beispielregister fuer Governance-, Audit- und Supplier-Kontexte mit Review- und Evidenzlogik.",
        width=6,
    )
    style_section_row(overview, 4, "Worum es in diesem Workbook geht", width=6)
    overview["A5"] = "Nutzen"
    overview["B5"] = "Dieses Workbook zeigt, wie ein dokumentationsreifes KI-Register aufgebaut sein sollte: Falllogik, Review-Status, Evidenzkette und naechster Review in derselben Struktur."
    overview["A6"] = "Geeignet fuer"
    overview["B6"] = "Governance-Teams, Berater, interne Auditoren, Lieferkettenanfragen"
    overview["A7"] = "Tabs"
    overview["B7"] = "Register | Review Matrix | Feldleitfaden"
    for row in overview.iter_rows(min_row=5, max_row=7, min_col=1, max_col=2):
        style_body_cell(row[0], bold=True, fill=SOFT)
        style_body_cell(row[1])
    set_widths(overview, {"A": 24, "B": 96, "C": 14, "D": 14, "E": 14, "F": 14})

    register = workbook.create_sheet("Register")
    set_sheet_base(register)
    style_title_block(
        register,
        "Register",
        "Beispielhafte Eintraege auf Einsatzfall-Ebene. Jeder Datensatz ist eine steuerbare Governance-Einheit.",
        width=12,
    )
    headers = [
        "Use Case ID",
        "Organisation",
        "Bereich",
        "Einsatzfall",
        "System",
        "Owner Rolle",
        "Datenbezug",
        "Entscheidungsnaehe",
        "AI-Act Risiko",
        "Review-Status",
        "Naechster Review",
        "Evidenz / Verify",
    ]
    register.append([])
    register.append(headers)
    style_table_header(register[4])
    rows = [
        [
            "UC-2026-001",
            "Nordstadt Automotive Systems GmbH",
            "Supplier Quality",
            "Automatisierte Eingangspruefung technischer Lieferantendokumente",
            "Technical Document Classifier",
            "Leitung Supplier Quality",
            "Technische Zeichnungen und Lastenhefte",
            "Vorbereitung mit manueller Freigabe",
            "Minimales Risiko",
            "REVIEWED",
            "2026-09-30",
            "Verify-Link + PDF Export",
        ],
        [
            "UC-2026-002",
            "Nordstadt Automotive Systems GmbH",
            "Einkauf",
            "Vorstrukturierung von Supplier Responses fuer OEM-Anfragen",
            "Supplier Request Assistant",
            "Leitung Einkauf",
            "Lieferantenstammdaten und Antworttexte",
            "Vorsortierung fuer menschliche Plausibilisierung",
            "Begrenztes Risiko",
            "PROOF_READY",
            "2026-10-15",
            "Use-Case Pass + Supplier Dossier",
        ],
        [
            "UC-2026-003",
            "Nordstadt Automotive Systems GmbH",
            "Legal & Governance",
            "Entwurf von Transparenzhinweisen fuer externe Kommunikation",
            "Disclosure Drafting Copilot",
            "Head of Governance",
            "Vertrags- und Kommunikationsentwuerfe",
            "Externe Kommunikation mit menschlicher Freigabe",
            "Begrenztes Risiko",
            "REVIEWED",
            "2026-08-31",
            "Review-Protokoll + Verify-Link",
        ],
        [
            "UC-2026-004",
            "Nordstadt Automotive Systems GmbH",
            "Operations",
            "KI-gestuetzte Priorisierung von Instandhaltungsauftraegen",
            "Maintenance Priority Engine",
            "Leitung Maintenance",
            "Maschinendaten und Stoerungsmeldungen",
            "Empfehlung fuer Einsatzplanung",
            "Begrenztes Risiko",
            "REVIEW_RECOMMENDED",
            "2026-05-31",
            "Review offen",
        ],
        [
            "UC-2026-005",
            "Nordstadt Automotive Systems GmbH",
            "HR",
            "Automatisierte Vorauswahl interner Weiterbildungsanfragen",
            "Learning Routing Assistant",
            "People Operations Lead",
            "Mitarbeiterdaten und Skill-Profile",
            "Priorisierung mit manueller Freigabe",
            "Begrenztes Risiko",
            "UNREVIEWED",
            "2026-06-15",
            "Erfassung erfolgt",
        ],
    ]
    for row in rows:
        register.append(row)
    for row in register.iter_rows(min_row=5, max_row=4 + len(rows), min_col=1, max_col=len(headers)):
        for cell in row:
            style_body_cell(cell)
        status_cell = row[9]
        status_cell.fill = status_fill(status_cell.value)
        status_cell.font = status_font(status_cell.value)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
    register.freeze_panes = "A5"
    add_excel_table(register, "RegisterMatrix", 4, 4 + len(rows), "L")
    set_widths(
        register,
        {
            "A": 14,
            "B": 28,
            "C": 18,
            "D": 34,
            "E": 24,
            "F": 20,
            "G": 26,
            "H": 24,
            "I": 18,
            "J": 18,
            "K": 14,
            "L": 24,
        },
    )

    review = workbook.create_sheet("Review Matrix")
    set_sheet_base(review)
    style_title_block(
        review,
        "Review Matrix",
        "Welche Fragen bei welchem Fall gestellt werden und welche Evidenz fuer eine Freigabe noetig ist.",
        width=7,
    )
    review.append([])
    review_headers = ["Use Case ID", "Trigger", "Pflicht-Reviewer", "Leitfragen", "Erforderliche Evidenz", "Offene Luecke", "Naechste Aktion"]
    review.append(review_headers)
    style_table_header(review[4])
    review_rows = [
        [
            "UC-2026-001",
            "Externe Lieferantendokumente werden vorsortiert",
            "Governance Lead",
            "Bleibt die finale Freigabe menschlich? Sind Dokumenttypen sauber begrenzt?",
            "Pass, Review-Protokoll, Kontrollbeschreibung",
            "Keine",
            "Quartalsreview halten",
        ],
        [
            "UC-2026-002",
            "Supplier Response beeinflusst Beschaffungsprozess",
            "Einkauf + Governance",
            "Wird nur vorstrukturiert oder bereits faktisch entschieden?",
            "Dossier, Request-Matrix, Sign-off",
            "FAQ fuer OEM-Rueckfragen ausbauen",
            "Response-Template finalisieren",
        ],
        [
            "UC-2026-004",
            "Priorisierung kann Instandhaltungsvorgaenge verschieben",
            "Operations + Governance",
            "Welche Auswirkungen hat ein Fehlranking? Gibt es Eskalation fuer kritische Anlagen?",
            "Review-Protokoll, Testfaelle, Owner-Freigabe",
            "Kritikalitaetsgrenze noch nicht dokumentiert",
            "Review terminieren",
        ],
    ]
    for row in review_rows:
        review.append(row)
    for row in review.iter_rows(min_row=5, max_row=4 + len(review_rows), min_col=1, max_col=len(review_headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(review, "ReviewQuestions", 4, 4 + len(review_rows), "G")
    set_widths(review, {"A": 14, "B": 24, "C": 18, "D": 30, "E": 30, "F": 24, "G": 22})

    guide = workbook.create_sheet("Feldleitfaden")
    set_sheet_base(guide)
    style_title_block(
        guide,
        "Feldleitfaden",
        "Kurze Begruendung der wichtigsten Registerfelder fuer Teams, die ein export- oder auditfaehiges Format aufbauen.",
        width=4,
    )
    guide.append([])
    guide_headers = ["Feld", "Warum es zaehlt", "Typischer Fehler", "Besserer Standard"]
    guide.append(guide_headers)
    style_table_header(guide[4])
    guide_rows = [
        ("Einsatzfall", "Die Nutzung ist die Governance-Einheit, nicht nur das Tool.", "Toolnamen ohne Kontext", "Ein konkreter Prozess pro Zeile"),
        ("Owner Rolle", "Ohne Verantwortung keine Review-Kette.", "Abteilung statt Rolle", "Benannte Funktion mit Freigabeverantwortung"),
        ("Entscheidungsnaehe", "Bestimmt den Pruefbedarf und die Eskalationslogik.", "Nur Nutzen, keine Wirkung", "Beschreiben, ob empfohlen, priorisiert oder entschieden wird"),
        ("Evidenz / Verify", "Ermoeglicht externe Anschlussfaehigkeit.", "Dateipfad ohne Systembezug", "Verify-Link oder geregelter Export"),
    ]
    for row in guide_rows:
        guide.append(row)
    for row in guide.iter_rows(min_row=5, max_row=4 + len(guide_rows), min_col=1, max_col=len(guide_headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(guide, "FieldGuide", 4, 4 + len(guide_rows), "D")
    set_widths(guide, {"A": 22, "B": 28, "C": 24, "D": 28})

    workbook.save(path)


def create_practitioner_workbook(path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    set_sheet_base(overview)
    style_title_block(
        overview,
        "Practitioner Workshop Kit",
        "Arbeitsmappe fuer Berater, Auditoren und Trainer, die einen Governance-Block produktiv in Workshops verankern wollen.",
        width=6,
    )
    style_section_row(overview, 4, "Empfohlener Zuschnitt", width=6)
    rows = [
        ("Einsatz", "Halbtag, Ganztag oder Governance Sprint"),
        ("Primarziel", "Reale KI-Einsatzfaelle sichtbar machen und in ein belastbares Register ueberfuehren"),
        ("Mindestergebnis", "Mindestens ein dokumentierter Einsatzfall, ein benannter Owner, ein Review-Pfad"),
        ("Beste Folgefrage", "Wer pflegt das Register weiter und wann ist der naechste Review?"),
    ]
    current_row = 5
    for label, text in rows:
        overview[f"A{current_row}"] = label
        overview[f"B{current_row}"] = text
        style_body_cell(overview[f"A{current_row}"], bold=True, fill=SOFT)
        style_body_cell(overview[f"B{current_row}"])
        current_row += 1
    set_widths(overview, {"A": 22, "B": 92, "C": 14, "D": 14, "E": 14, "F": 14})

    agenda = workbook.create_sheet("Workshop Canvas")
    set_sheet_base(agenda)
    style_title_block(
        agenda,
        "Workshop Canvas",
        "Nutzen Sie diese Struktur, um operative KI-Arbeit direkt in dokumentierbare Governance zu ueberfuehren.",
        width=6,
    )
    agenda.append([])
    headers = ["Phase", "Leitfrage", "Moderationshinweis", "Output", "Owner", "Zeit"]
    agenda.append(headers)
    style_table_header(agenda[4])
    rows = [
        ("Einstieg", "Welche KI-Systeme werden heute real genutzt?", "Erst Breite sammeln, noch nicht bewerten.", "Rohinventar", "Moderator", "20 Min"),
        ("Use-Case Scharfstellung", "Welche Einsaetze sind fuer Kunden, Mitarbeiter oder Entscheidungen relevant?", "Toollisten in konkrete Einsaetze uebersetzen.", "3-5 priorisierte Faelle", "Fachbereich", "35 Min"),
        ("Governance-Block", "Wer verantwortet den Fall und wie wird er freigegeben?", "Ownership und Review nicht vertagen.", "Owner + Review-Pfad", "Governance", "30 Min"),
        ("Capture", "Welche Felder werden sofort im Register befuellt?", "Mindestens einen echten Datensatz live anlegen.", "Erster Pass / erster Registereintrag", "Teamlead", "25 Min"),
        ("Abschluss", "Welche Folgeaktion ist vor dem naechsten Termin faellig?", "Naechsten Review konkret terminieren.", "Action List", "Projektverantwortung", "10 Min"),
    ]
    for row in rows:
        agenda.append(row)
    for row in agenda.iter_rows(min_row=5, max_row=4 + len(rows), min_col=1, max_col=len(headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(agenda, "WorkshopAgenda", 4, 4 + len(rows), "F")
    set_widths(agenda, {"A": 18, "B": 24, "C": 30, "D": 22, "E": 18, "F": 10})

    capture = workbook.create_sheet("Capture Template")
    set_sheet_base(capture)
    style_title_block(
        capture,
        "Capture Template",
        "Live-Tabelle fuer die ersten erfassten Einsatzfaelle waehrend Workshop oder Governance Sprint.",
        width=10,
    )
    capture.append([])
    capture_headers = [
        "Team / Bereich",
        "Einsatzfall",
        "System",
        "Nutzungskontext",
        "Owner Rolle",
        "Datenbezug",
        "Entscheidungsnaehe",
        "AI-Act Risiko",
        "Review noetig?",
        "Naechster Schritt",
    ]
    capture.append(capture_headers)
    style_table_header(capture[4])
    example_rows = [
        ("Einkauf", "Vorstrukturierung von OEM-Rueckfragen", "Supplier Request Assistant", "Beschaffungsnahe Aufbereitung", "Leitung Einkauf", "Lieferanten- und Dokumentationsdaten", "Vorbereitung", "Begrenzt", "Ja", "Owner und Freigaberegel festhalten"),
        ("Qualitaet", "Pruefung technischer Dokumente", "Technical Document Classifier", "Vorpruefung vor manueller Freigabe", "Leitung Supplier Quality", "Technische Lieferantendokumente", "Vorbereitung", "Minimal", "Ja", "Review-Zyklus setzen"),
    ]
    for row in example_rows:
        capture.append(row)
    for row in capture.iter_rows(min_row=5, max_row=14, min_col=1, max_col=len(capture_headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(capture, "CaptureTemplate", 4, 14, "J")
    set_widths(capture, {"A": 18, "B": 28, "C": 22, "D": 28, "E": 18, "F": 24, "G": 20, "H": 14, "I": 14, "J": 24})

    checklist = workbook.create_sheet("Delivery Checklist")
    set_sheet_base(checklist)
    style_title_block(
        checklist,
        "Delivery Checklist",
        "Was nach dem Termin vorliegen sollte, damit das Ergebnis fuer den Kunden mehr ist als nur eine gute Diskussion.",
        width=4,
    )
    checklist.append([])
    checklist_headers = ["Checkpunkt", "Warum wichtig", "Erledigt", "Kommentar"]
    checklist.append(checklist_headers)
    style_table_header(checklist[4])
    checklist_rows = [
        ("Mindestens ein realer Einsatzfall dokumentiert", "Der Kunde braucht ein Artefakt, kein nur gesprochenes Ergebnis.", "", ""),
        ("Owner pro priorisiertem Einsatzfall benannt", "Ohne Owner gibt es keine fortfuehrbare Governance.", "", ""),
        ("Naechster Review terminlich fixiert", "Follow-up darf nicht implizit bleiben.", "", ""),
        ("Download oder Verify-Bezug erklaert", "Der Kunde muss verstehen, wie externe Nachweise spaeter entstehen.", "", ""),
        ("Folgeauftrag oder Betreuungspfad benannt", "Governance lebt von Weiterpflege, nicht von Einmalsichtung.", "", ""),
    ]
    for row in checklist_rows:
        checklist.append(row)
    for row in checklist.iter_rows(min_row=5, max_row=4 + len(checklist_rows), min_col=1, max_col=len(checklist_headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(checklist, "DeliveryChecklist", 4, 4 + len(checklist_rows), "D")
    set_widths(checklist, {"A": 34, "B": 42, "C": 12, "D": 24})

    workbook.save(path)


def create_supplier_request_workbook(path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    set_sheet_base(overview)
    style_title_block(
        overview,
        "Supplier Request & Response Kit",
        "Vorlage fuer strukturierte OEM- oder Partneranfragen zu einem einzelnen KI-Einsatzfall inklusive Beispielantwort und Evidenzliste.",
        width=6,
    )
    style_section_row(overview, 4, "Wofuer dieses Workbook gedacht ist", width=6)
    rows = [
        ("Ziel", "Rueckfragen zu einem konkreten KI-Einsatzfall standardisiert beantworten"),
        ("Vermeidet", "Freiform-E-Mails, doppelte Pflege, widerspruechliche Lieferantenaussagen"),
        ("Beste Nutzung", "Beschaffung, Lieferkette, Partnerpruefung, OEM-Governance"),
        ("Immer koppeln mit", "Use-Case Pass, Review-Stand, Verify-Link oder kontrolliertem Export"),
    ]
    current_row = 5
    for label, text in rows:
        overview[f"A{current_row}"] = label
        overview[f"B{current_row}"] = text
        style_body_cell(overview[f"A{current_row}"], bold=True, fill=SOFT)
        style_body_cell(overview[f"B{current_row}"])
        current_row += 1
    set_widths(overview, {"A": 22, "B": 92, "C": 14, "D": 14, "E": 14, "F": 14})

    request = workbook.create_sheet("Request Template")
    set_sheet_base(request)
    style_title_block(
        request,
        "Request Template",
        "Blanko-Vorlage fuer strukturierte Lieferkettenanfragen. Jede Zeile bezieht sich auf einen Punkt, der fuer die Antwort und spaetere Nachweisfuehrung zaehlt.",
        width=4,
    )
    request.append([])
    headers = ["Feld", "Pflicht", "Warum der Auftraggeber es wissen will", "Antwort Lieferant"]
    request.append(headers)
    style_table_header(request[4])
    rows = [
        ("Lieferantenorganisation", "Ja", "Juristische Zuordnung des antwortenden Unternehmens", ""),
        ("Einsatzfall / Titel", "Ja", "Es wird ein einzelner Fall und nicht die gesamte Tool-Landschaft angefragt", ""),
        ("Zweck", "Ja", "Beschreibung der realen Nutzung im Prozess", ""),
        ("Owner Rolle", "Ja", "Verantwortung und Rueckfragemoeglichkeit", ""),
        ("System / Modell", "Ja", "Technische Einordnung und Produktbezug", ""),
        ("Datenbezug", "Ja", "Welche Datenarten oder Dokumente betroffen sind", ""),
        ("Entscheidungsnaehe", "Ja", "Ob vorbereitet, priorisiert oder final entschieden wird", ""),
        ("Externe Wirkung", "Ja", "Ob OEM, Kunden oder Mitarbeitende indirekt betroffen sind", ""),
        ("Review-Status", "Ja", "Wie weit der Fall intern geprueft wurde", ""),
        ("Verify / Artefakt Link", "Nein", "Wie die Aussage extern bestaetigt werden kann", ""),
        ("Offene Luecken", "Nein", "Welche Punkte bewusst noch nicht als final bestaetigt werden", ""),
    ]
    for row in rows:
        request.append(row)
    for row in request.iter_rows(min_row=5, max_row=4 + len(rows), min_col=1, max_col=len(headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(request, "SupplierRequest", 4, 4 + len(rows), "D")
    set_widths(request, {"A": 24, "B": 10, "C": 42, "D": 34})

    response = workbook.create_sheet("Example Response")
    set_sheet_base(response)
    style_title_block(
        response,
        "Example Response",
        "Beispielhaft ausgefuellte Antwort fuer einen dokumentierten Automotive-Einsatzfall.",
        width=4,
    )
    response.append([])
    response.append(headers)
    style_table_header(response[4])
    sample_rows = [
        ("Lieferantenorganisation", "Ja", "Juristische Zuordnung des antwortenden Unternehmens", "Nordstadt Automotive Systems GmbH"),
        ("Einsatzfall / Titel", "Ja", "Es wird ein einzelner Fall und nicht die gesamte Tool-Landschaft angefragt", "Automatisierte Eingangspruefung technischer Lieferantendokumente"),
        ("Zweck", "Ja", "Beschreibung der realen Nutzung im Prozess", "Vorpruefung technischer Unterlagen vor manueller Freigabe"),
        ("Owner Rolle", "Ja", "Verantwortung und Rueckfragemoeglichkeit", "Leitung Supplier Quality"),
        ("System / Modell", "Ja", "Technische Einordnung und Produktbezug", "Technical Document Classifier"),
        ("Datenbezug", "Ja", "Welche Datenarten oder Dokumente betroffen sind", "Technische Zeichnungen, Lastenhefte, Lieferantendokumente"),
        ("Entscheidungsnaehe", "Ja", "Ob vorbereitet, priorisiert oder final entschieden wird", "Vorbereitung mit manueller Freigabe"),
        ("Externe Wirkung", "Ja", "Ob OEM, Kunden oder Mitarbeitende indirekt betroffen sind", "Indirekte Wirkung auf Beschaffungs- und Qualitaetsprozess"),
        ("Review-Status", "Ja", "Wie weit der Fall intern geprueft wurde", "REVIEWED"),
        ("Verify / Artefakt Link", "Nein", "Wie die Aussage extern bestaetigt werden kann", "Verify-Link + PDF-Export des Passes"),
        ("Offene Luecken", "Nein", "Welche Punkte bewusst noch nicht als final bestaetigt werden", "Keine finale Freigabe ohne menschliche Sichtung"),
    ]
    for row in sample_rows:
        response.append(row)
    for row in response.iter_rows(min_row=5, max_row=4 + len(sample_rows), min_col=1, max_col=len(headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(response, "SupplierResponse", 4, 4 + len(sample_rows), "D")
    set_widths(response, {"A": 24, "B": 10, "C": 42, "D": 34})

    evidence = workbook.create_sheet("Evidence Checklist")
    set_sheet_base(evidence)
    style_title_block(
        evidence,
        "Evidence Checklist",
        "Pruefpunkte fuer Antworten, die belastbar sein sollen und nicht bei Rueckfragen zerfallen duerfen.",
        width=5,
    )
    evidence.append([])
    evidence_headers = ["Beleg", "Verantwortung", "Fuer wen relevant", "Status", "Kommentar"]
    evidence.append(evidence_headers)
    style_table_header(evidence[4])
    evidence_rows = [
        ("Use-Case Pass", "Governance Lead", "OEM / Procurement", "Vorhanden", "Pass ist der Primarbeleg"),
        ("Review-Protokoll", "Governance Lead", "Audit / OEM", "Vorhanden", "Letzter Review dokumentiert"),
        ("Verify-Link", "Register Owner", "Externe Echtheitspruefung", "Vorhanden", "Nur fuer freigegebene Faelle"),
        ("Offene Punkte", "Owner Rolle", "Interne Nachverfolgung", "Pflegen", "Keine improvisierten E-Mail-Antworten"),
    ]
    for row in evidence_rows:
        evidence.append(row)
    for row in evidence.iter_rows(min_row=5, max_row=4 + len(evidence_rows), min_col=1, max_col=len(evidence_headers)):
        for cell in row:
            style_body_cell(cell)
    add_excel_table(evidence, "EvidenceChecklist", 4, 4 + len(evidence_rows), "E")
    set_widths(evidence, {"A": 24, "B": 20, "C": 24, "D": 14, "E": 26})

    workbook.save(path)


def generate_assets() -> None:
    ensure_dirs()
    sync_handouts()
    create_use_case_pass_pdf(EXAMPLE_DIR / "ki-register-use-case-pass-beispiel.pdf")
    create_supplier_response_pdf(EXAMPLE_DIR / "ki-register-supplier-response-beispiel.pdf")
    create_use_case_register_workbook(WORKBOOK_DIR / "ki-register-use-case-register-beispiel.xlsx")
    create_practitioner_workbook(WORKBOOK_DIR / "ki-register-practitioner-workshop-kit.xlsx")
    create_supplier_request_workbook(WORKBOOK_DIR / "ki-register-supplier-request-template.xlsx")


if __name__ == "__main__":
    generate_assets()
